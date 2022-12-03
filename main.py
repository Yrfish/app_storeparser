from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re
import time
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from fake_useragent import UserAgent

wb = load_workbook(filename="app_store_sheet.xlsx", data_only=True)
ws = wb.active
urls = list()
index_list = list()
for row in range(3, ws.max_row+1):
    if ws["F" + f"{row}"].value is None:
        url = ws["E" + f"{row}"].value
        urls.append(url)
        index_list.append(row)


def crawl(urls, index_list):
    i = 0
    lost_email = 0
    index = 0
    for url in urls:

        proxies = {
            "103.189.235.198": "3128",
            "113.53.61.163": "8080",
            "138.118.104.166": "999",
            "52.81.149.245": "3128",
            "159.192.138.170": "	8080",
            "92.101.95.210": "1080",
            "27.79.10.160": "10000",
            "157.245.207.186": "8080",
            "167.99.201.165": "32295",
            "167.172.173.210": "39377",
            "177.141.99.50": "8080"

        }

        ua = UserAgent(browsers=["chrome"]).random
        headers = {
            "user-agent":ua
        }

        print(f"\nnow i parse url - {url}")

        print(f"headers =  {headers}")
        rqst = requests.get(url=url, headers=headers, proxies=proxies)
        print(f"response = {rqst}")
        src = rqst.text

        # Save current url into local machine it save us form blocks
        with open("index.html", "w", encoding="utf-8") as file:
            file.write(src)
        # open file in file handle "page"
        with open("index.html", encoding="utf-8") as f:
            page = f.read()

        # Start collecting data about app

        soup = BeautifulSoup(page, "lxml")
# try block that check if subtitile exist. If it does not, it can:
        # Bad url response 404, too much requests and response 429
        try:
            subtitle = soup.find("h2").text
# except runs if try are wrong and check if response 404 it assing cant connect to F cell, or if response 429 wait 40
        # secs and try again
        except:
            if rqst.status_code == 404:
                ws[f"F{index_list[index]}"] = "cant connect"
                wb.save("app_store_sheet.xlsx")
                print(f"response {rqst} -- {url}")
                index += 1
                continue
            elif rqst.status_code == 429:
                print(f"response -- {rqst} -- wait 40sec and try again")
                time.sleep(40)
                rqst = requests.get(url=url, headers=headers, proxies=proxies)


            src = rqst.text

                # Save current url into local machine it save us form blocks
            with open("index.html", "w", encoding="utf-8") as file:
                file.write(src)
                # open file in file handle "page"
            with open("index.html", encoding="utf-8") as f:
                page = f.read()

            soup = BeautifulSoup(page, "lxml")
            subtitle = soup.find("h2").text

        description = soup.find(
            class_=
            "we-truncate we-truncate--multi-line we-truncate--interactive l-column small-12 medium-9 large-8").text

        dev_url_lst = soup.find(class_="link icon icon-after icon-external").get_attribute_list("href")
        dev_url = ",".join(dev_url_lst)
        p_tags = soup.find_all("p")

        # Going through p_tag anf searching for p tag that contains English. I took English cos it is the most frequent language
        for p_tag in p_tags:
            if "English" or "Englisch" or " Anglais" in p_tag.text and len(p_tag.text) < 600:
                languages = p_tag.text

        try:
            rating = soup.find(class_="we-customer-ratings__averages__display").text
            rating_score2words = soup.find(class_="we-customer-ratings__count small-hide medium-show").text
            # Remove word "Ratings" from string
            rating_score = rating_score2words[:-(rating_score2words.endswith("Ratings") and len("Ratings"))]
        except:
            print("app has no rating")
            rating_score = "N.A."
            rating = "N.A."

        copyright = soup.find(
            class_="information-list__item__definition information-list__item__definition--copyright").text

        try:
            privacy_url = "".join(soup.find("a", {'data-metrics-click': re.compile(
                '{"actionType":"navigate","targetType":"link","targetId":"LinkToPrivacyPolicy"}')}).get_attribute_list(
                "href"))
            print(f"privacy site {privacy_url}")
        except:
            privacy_url = "N.A"

        try:
            app_sup = "".join(soup.find("a", {'data-metrics-click': '{"actionType":"navigate","targetType":"link","targetId":"LinkToAppSupport"}'}).get_attribute_list("href"))
        except:
            app_sup = "N.A."

        #  ↑ ↑ ↑ ↑ ↑ ↑ ↑ above i collect all info on AppStore ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑

        # Start searching for email at privacy pol page
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        try:
            driver.get(
                privacy_url)  # try to go to privacy policy page. As i mentioned it can be pdf file or does not exist
            time.sleep(3)
            privacy_page_code = driver.page_source
            email = ", ".join(set(re.findall(r"[A-Za-z0-9._-]{1,15}@(?!png)[\w-]{2,}\.?[A-Za-z.]+",
                                             privacy_page_code)))  # Get list of emails there where must be 1 or more character befor "@" and it include only one dot
            # Checking if it not pdf but cant find email
            if not email:
                email = "N.A"
                lost_email += 1
                print(f"there is the psge where i cant find email {privacy_url}")
            print(f"it is a comoany email {email}")

        except:
            email = "None"
            print(privacy_url, "does not exist")

        ws[f"F{index_list[index]}"] = description
        ws[f"G{index_list[index]}"] = subtitle
        ws[f"H{index_list[index]}"] = dev_url
        ws[f"I{index_list[index]}"] = app_sup
        ws[f"J{index_list[index]}"] = privacy_url
        ws[f"K{index_list[index]}"] = email
        ws[f"L{index_list[index]}"] = "N.A"
        ws[f"M{index_list[index]}"] = "N.A"
        ws[f"N{index_list[index]}"] = rating
        ws[f"O{index_list[index]}"] = rating_score
        ws[f"P{index_list[index]}"] = copyright
        ws[f"Q{index_list[index]}"] = languages
        wb.save("app_store_sheet.xlsx")

        index += 1
        i += 1
        if i == 25:
            i = 0
            print("Saving the file app_store_sheet.xlsx to current directory")
            wb.save("app_store_sheet.xlsx")
            print("\nsnooooozee 30sec >>>> zzz....zz..zzzz.zz")
            time.sleep(30)
            print("\n i woke up and continue working")


crawl(urls, index_list)
